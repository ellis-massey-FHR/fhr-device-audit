# scripts/get_servicenow_data.py
import os
from io import BytesIO
from pathlib import Path

import requests
from requests.auth import HTTPBasicAuth
from dotenv import load_dotenv
import pandas as pd

# ---- Headless matplotlib for CI/Codespaces BEFORE pyplot import ----
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage

# -----------------------Load environment----------------------
env_path = Path(__file__).resolve().parent / "servicenow.env"
load_dotenv(dotenv_path=env_path)

# Allow fallback to CI secrets if .env is absent
instance = os.getenv("SERVICENOW_INSTANCE") or os.environ.get("SN_INSTANCE")
username = os.getenv("SERVICENOW_USER") or os.environ.get("SN_USER")
password = os.getenv("SERVICENOW_PASS") or os.environ.get("SN_PASS")

print("ENV path:", env_path)
print("Instance loaded:", instance)
print("Username loaded:", username)
print("Password loaded:", bool(password))

# Output directory: default to ./output in CI, can be overridden via env
OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", "output"))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_PATH = OUTPUT_DIR / "fhr_computer_requests.xlsx"

#-----------------------Helpers-----------------------
def fetch_details(link):
    if not link:
        return {}
    res = requests.get(
        link,
        auth=HTTPBasicAuth(username, password),
        params={"sysparm_display_value": "true"},
        timeout=30,
    )
    if res.status_code == 200:
        return res.json().get("result", {})
    return {}

def format_workstation_string(ws_string):
    if not isinstance(ws_string, str):
        return ""
    # Normalize escaped newlines, trim commas/spaces
    cleaned = ws_string.replace("\\r", "").replace("\\n", ", ").replace("\r", " ").replace("\n", " ")
    return cleaned.strip().lstrip(", ")

def get_current_workstation_from_ci(user_sys_id):
    if not user_sys_id:
        return None

    ci_url = f"{instance}/api/now/table/cmdb_ci_computer"
    params = {
        "sysparm_query": (
            f"assigned_to={user_sys_id}"
            "^model_category.name=Computer"
            "^ORDERBYDESClast_discovered"
        ),
        "sysparm_display_value": "true",
        "sysparm_limit": "50",
    }

    response = requests.get(ci_url, auth=HTTPBasicAuth(username, password), params=params, timeout=60)
    if response.status_code != 200:
        print(f"❌ Error fetching CIs: {response.status_code}")
        return None

    records = response.json().get("result", [])
    all_devices = []

    for r in records:
        name = r.get("name")
        class_name = r.get("sys_class_name")

        # OS handling: include Windows or UNKNOWN; skip only if we know it's not Windows
        os_raw = r.get("os")
        os_name = os_raw.lower() if isinstance(os_raw, str) else ""

        if not os_name:
            alt_os = r.get("operating_system") or r.get("u_os_full_name")
            if isinstance(alt_os, str):
                os_name = alt_os.lower()
            elif isinstance(alt_os, dict):
                os_name = (alt_os.get("display_value") or alt_os.get("value") or "").lower()

        if os_name and "windows" not in os_name:
            # Example: iPhones etc.
            printable = name.replace("\n", " ").replace("\r", " ").strip() if isinstance(name, str) else name
            print(f"🛑 Skipping non-Windows device: {printable} (OS: {os_name})")
            continue

        if isinstance(name, str):
            name = name.replace("\n", " ").replace("\r", " ").strip()

        if name and class_name and name not in all_devices:
            all_devices.append(name)

    return format_workstation_string(", ".join(all_devices)) if all_devices else None

#----------------------Query ServiceNow--------------------------
url = f"{instance}/api/now/table/sc_req_item"
params = {
    "sysparm_query": (
        "company.name=Flint Hills Resources"
        "^u_new_hire=false"
        "^cat_item.nameLIKElaptop"
        "^ORcat_item.nameLIKEsurface"
        "^opened_atRELGTjavascript:gs.daysAgoStart(60)^ORclosed_atRELGTjavascript:gs.daysAgoStart(60)"
        "^ORDERBYDESCclosed_at"
    ),
    "sysparm_display_value": "all",
    "sysparm_limit": "115",
}

response = requests.get(url, auth=HTTPBasicAuth(username, password), params=params, timeout=120)

if response.status_code != 200:
    print(f"❌ Error: {response.status_code}")
    print(response.text)
    raise SystemExit(1)

print("✅ Success!")
data = response.json()
results = data.get("result", [])
print(f"Total records returned: {len(results)}")

rows = []
for item in results:
    print("→ Processing item ID:", item.get("sys_id"))

    requested_for_info = fetch_details(item.get("requested_for", {}).get("link"))

    user_location = (
        requested_for_info.get("location", {}).get("display_value")
        or requested_for_info.get("u_location", {}).get("display_value")
        or requested_for_info.get("u_site_id")
        or "Unknown"
    )

    user_sys_id = requested_for_info.get("sys_id")
    recent_ws = get_current_workstation_from_ci(user_sys_id)

    if not recent_ws:
        recent_ws = format_workstation_string(requested_for_info.get("u_workstations", ""))

    rows.append({
        "Requested For": item.get("requested_for", {}).get("display_value", "Unknown"),
        "Requested For Email": requested_for_info.get("email", "Unknown"),
        "Requested For Location": user_location,
        "Catalog Item": item.get("cat_item", {}).get("display_value", "Unknown"),
        "Created": item.get("opened_at", {}).get("display_value", "Unknown") if isinstance(item.get("opened_at"), dict) else item.get("opened_at", "Unknown"),
        "Closed": item.get("closed_at", {}).get("display_value", "Unknown") if isinstance(item.get("closed_at"), dict) else item.get("closed_at", "Unknown"),
        "Workstations": recent_ws or "Unknown",
    })

df = pd.DataFrame(rows)

#------------------------------Group by Location--------------------
def normalize_location(loc):
    if not loc:
        return "Unknown"
    loc_up = str(loc).upper()
    if "CORPUSCHRISTI" in loc_up or "CORPUS CHRISTI" in loc_up:
        return "Corpus Christi"
    elif "ROSEMOUNT" in loc_up or "PINE BEND" in loc_up:
        return "Rosemount"
    elif "WICHITA" in loc_up:
        return "Wichita"
    else:
        return "Other"

df["Location Group"] = df["Requested For Location"].apply(normalize_location)

#--------------------------Write to Excel (must happen BEFORE load_workbook)-------------------------
df.to_excel(EXCEL_PATH, index=False, engine="openpyxl")
print(f"📁 Wrote base Excel to {EXCEL_PATH}")

#---------------------------Open and format workbook-------------------------
wb = load_workbook(EXCEL_PATH)
ws = wb.active

# Insert title row and style it
ws.insert_rows(1)
header_text = "FHR Laptop Requests Report"
total_columns = ws.max_column
end_col_letter = get_column_letter(total_columns)
ws.merge_cells(f"A1:{end_col_letter}1")
header_cell = ws["A1"]
header_cell.value = header_text
header_cell.font = Font(size=16, bold=True)
header_cell.alignment = Alignment(horizontal="center", vertical="center")

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column
    col_letter = get_column_letter(column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except Exception:
            pass
    if ws.cell(row=2, column=column).value == "Workstations":
        ws.column_dimensions[col_letter].width = 40
    else:
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

#------------------------Embed pie chart--------------------------
location_counts = df["Location Group"].value_counts()

def make_autopct(values):
    def my_autopct(pct):
        total = sum(values)
        count = int(round(pct * total / 100.0))
        return f"{count}"
    return my_autopct

pie_buffer = BytesIO()
plt.figure(figsize=(6, 6))
plt.pie(location_counts, labels=location_counts.index, autopct=make_autopct(location_counts), startangle=140)
plt.title(f"Laptop Requests by Location (Total: {location_counts.sum()})")
plt.axis("equal")
plt.tight_layout()
plt.savefig(pie_buffer, format="png")
plt.close()

pie_buffer.seek(0)
img = XLImage(pie_buffer)
img.width = 350
img.height = 350
ws.add_image(img, "J6")
print("📊 Pie chart embedded in Excel at J6")

wb.save(EXCEL_PATH)
print(f"✅ Excel with embedded chart saved to {EXCEL_PATH}")
